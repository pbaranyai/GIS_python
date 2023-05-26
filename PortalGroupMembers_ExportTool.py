# -*- coding: utf-8 -*-
# ---------------------------------------------------------------------------
# PortalGroupMembers_ExportTool.py
# Created on: 2023-05-25
# Updated on: 2023-05-25
# Works in ArcGIS Pro
#
# Author: Phil Baranyai/DLC
#
# Description:
# List all groups in portal/AGOL and provide output list of groups and members within.
#
#
# Works with Enterprise GIS & ArcGIS Online
#
# ---------------------------------------------------------------------------
print("This tool reads from portal URL entered below, lists all groups and members within by username, and export out results an excel report.")
print("\nLoading python modules, please wait...")
from arcgis.gis import GIS
import pandas as pd
import os
import datetime
from openpyxl import load_workbook

# Comment out for manual run of script - Used for prompts within command window (Run with ArcGIS Pro)
print("Enter portal URL below: | Example: https://PORTALNAME.com/arcgis")
Portal = input('Enter URL here: ')
PortalUserName = input('Enter username for '+str(Portal)+'here: ')

# This can be used in leiu of CMD window, need to comment out CMD window portions of script at top and bottom, then un-comment out Portal variable here. 
#*********************************************************************************************************
##### Change portal address here  ##### 
#Portal = 'https://PORTALNAME.com/arcgis' <-- Used only for manual run of script
##### Change portal address here  #####
#*********************************************************************************************************

# Setup Date/time variables
date = datetime.date.today().strftime("%Y%m%d")
Day = time.strftime("%Y-%m-%d", time.localtime())
Time = time.strftime("%H%M", time.localtime())
start_time = time.time()
elapsed_time = time.time() - start_time

# Setup export path to *script location* PortalGroupMembers_Reports folder
ReportDirectory = os.getcwd()+"\\PortalGroupMembers_Reports"
reportdirExists = os.path.exists(ReportDirectory)
if not reportdirExists:
    os.makedirs(ReportDirectory)
    print(ReportDirectory+" was not found, so it was created")

# Confirm portal access was successful for Portal
for url in Portal:
    print("Attempting login on: "+str(url)+" | Password required for "+str(PortalUserName))
    gis = GIS(url,PortalUserName)
    LoggedInAs = gis.properties.user.username
    # Clean up Portal url for usable name in print statements and excel file name
    PortalName = Portal.replace('https://','',1).replace('.com/arcgis','',1)
    print("\nLogged into "+str(PortalName)+" as "+str(LoggedInAs)+" at "+str(Time)+" hrs, beginning report")
else:
    print("Unable to login to "+str(Portal)+", check portal URL & password (if applicable) and try again, if portal URL is correct, ensure you have proper access via your login credentials")

# Set Excel spreadsheet output name
ExcelOutput = os.path.join(ReportDirectory,str(PortalName)+'__Portal_Group_Members_report__'+str(date)+"_"+str(Time)+'.xlsx')

#Create empty list to store group members within
group_members_list = []

# Establish a variable called groups (iterating through the established portal, and listing groups)
groups = gis.groups.search('*', max_groups=1000)

# Iterate through groups list, establish variable called groupMembers (users from each group), append results to dictionary
for group in groups:
    groupMembers = group.get_members()['users']
    group_members_list.append({'Group Name':group.title,'Group Members':groupMembers})

# Create dataframe from dictionary entries
df = pd.DataFrame(group_members_list)

print("+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+")
# Exporting Dataframe to excel
print('\nExporting to Excel, located at: '+ExcelOutput)
df.to_excel(ExcelOutput, 'Groups', index=False)

# Access exported excel workbook, and auto-size columns for easier read
wb = load_workbook(ExcelOutput)
ws = wb['Groups']
for letter in ['A', 'B']:
    max_width = int(0)
    for row_number in range(1,ws.max_row +1):
        if len(ws[f'{letter}{row_number}'].value) > max_width:
               max_width = len(ws[f'{letter}{row_number}'].value)
    ws.column_dimensions[letter].width = max_width +1
wb.save(ExcelOutput)

# Calculating run time and printing end statement
end_time = time.strftime("%I:%M:%S %p", time.localtime())
elapsed_time = time.time() -start_time
print("\nReporting process completed at " + str(end_time)+" taking "+time.strftime("%M minutes %S seconds", time.gmtime(elapsed_time)))

# For command run, allows user to see all program print statements before closing the command window by pressing any key.
input("Press enter key to close program")
sys.exit()
